#!/usr/bin/env python3
"""
Pathology Report Batch Compliance Checker

Analyzes pathology reports against CAP/ICCR guidelines using Claude API
or local LLMs (Ollama, LM Studio).

Usage:
    python batch_checker.py <input_dir> <output_dir> [options]

Arguments:
    input_dir           Directory containing report text files (.txt)
    output_dir          Directory to save results

Options:
    --provider, -p      LLM provider: anthropic (default), ollama, lmstudio
    --model, -m         Model name (auto-detected for local providers)
    --tumor-type, -t    Tumor type: breast, colorectal, pancreas, gastric
    --list-models       List available models for the selected provider
    --ollama-url        Ollama API URL (default: http://localhost:11434/v1)
    --lmstudio-url      LM Studio API URL (default: http://localhost:1234/v1)

Environment:
    ANTHROPIC_API_KEY   Required for Claude provider
    LLM_PROVIDER        Default provider
    OLLAMA_MODEL        Default Ollama model
    LMSTUDIO_MODEL      Default LM Studio model

Examples:
    # With Claude (default)
    python batch_checker.py /reports /output

    # With local Ollama
    python batch_checker.py /reports /output -p ollama -m qwen2.5:3b

    # With LM Studio
    python batch_checker.py /reports /output -p lmstudio

Output files:
    - individual_reports/  Folder with per-report QA results
    - summary_report.txt   Overall statistics
    - compliance_data.csv  Structured data for all reports
    - compliance_data.xlsx Excel workbook with multiple sheets
    - trend_history.json   Historical tracking data
"""

import os
import sys
import json
import csv
import re
import time
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional, Tuple, Any

# Try to import optional packages
ANTHROPIC_AVAILABLE = False
OPENAI_AVAILABLE = False
EXCEL_AVAILABLE = False

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    pass

try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    print("Note: openpyxl not installed. Excel export will be skipped.")


# ============================================================================
# CONFIGURATION
# ============================================================================

DEFAULT_ANTHROPIC_MODEL = "claude-sonnet-4-20250514"
DEFAULT_OLLAMA_MODEL = "llama3.1:70b"
DEFAULT_OLLAMA_URL = "http://localhost:11434/v1"
DEFAULT_LMSTUDIO_URL = "http://localhost:1234/v1"


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class GapInfo:
    element: str
    severity: str  # "critical", "major", "minor"
    status: str    # "missing" or "empty"
    guideline: str
    required_for: str


@dataclass
class ValidationIssue:
    rule_name: str
    description: str
    found_value: str
    expected_value: str
    severity: str  # "error" or "warning"
    note: str


@dataclass
class QualityMetrics:
    completeness_score: float
    clarity_score: float
    consistency_score: float
    overall_quality: float
    completeness_details: str
    clarity_details: str
    consistency_details: str


@dataclass
class StagingInfo:
    pT: Optional[str] = None
    pN: Optional[str] = None
    pM: Optional[str] = None
    reported_stage: Optional[str] = None
    calculated_stage: Optional[str] = None
    stage_match: Optional[bool] = None
    stage_notes: str = ""


@dataclass
class ReportResult:
    filename: str
    tumor_type: str
    score: int
    status: str
    critical_count: int
    major_count: int
    minor_count: int
    missing_count: int
    empty_count: int
    present_count: int
    gaps: List[GapInfo] = field(default_factory=list)
    present_elements: List[str] = field(default_factory=list)
    validation_issues: List[ValidationIssue] = field(default_factory=list)
    quality_metrics: Optional[QualityMetrics] = None
    staging: Optional[StagingInfo] = None
    raw_analysis: str = ""


# ============================================================================
# MODEL DISCOVERY
# ============================================================================

def list_available_models(base_url: str, provider: str, quiet: bool = False) -> List[str]:
    """Query the /v1/models endpoint to list available models."""
    if not OPENAI_AVAILABLE:
        if not quiet:
            print("Warning: openai package not installed.", file=sys.stderr)
        return []

    try:
        client = OpenAI(base_url=base_url, api_key="not-needed")
        models_response = client.models.list()
        model_ids = [m.id for m in models_response.data]
        return model_ids
    except Exception as e:
        if not quiet:
            print(f"Warning: Could not list models from {base_url}: {e}", file=sys.stderr)
        return []


def auto_select_model(base_url: str, provider: str, quiet: bool = False) -> Optional[str]:
    """Auto-detect and select a model from the local server."""
    models = list_available_models(base_url, provider, quiet)

    if not models:
        return None

    if len(models) == 1:
        if not quiet:
            print(f"Auto-detected model: {models[0]}")
        return models[0]

    # Multiple models available
    if not quiet:
        print(f"\nMultiple models available in {provider}:")
        for i, model_id in enumerate(models, 1):
            print(f"  {i}. {model_id}")
        print(f"\nUsing first available model: {models[0]}")
        print(f"Tip: Use --model to select a specific model.\n")

    return models[0]


# ============================================================================
# LLM CLIENT CREATION
# ============================================================================

def create_anthropic_client():
    """Create Anthropic client."""
    if not ANTHROPIC_AVAILABLE:
        print("Error: anthropic package not installed.", file=sys.stderr)
        print("Install with: pip install anthropic", file=sys.stderr)
        sys.exit(1)

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("Error: ANTHROPIC_API_KEY environment variable not set.", file=sys.stderr)
        sys.exit(1)

    return anthropic.Anthropic(api_key=api_key)


def create_openai_compatible_client(base_url: str):
    """Create OpenAI-compatible client for Ollama or LM Studio."""
    if not OPENAI_AVAILABLE:
        print("Error: openai package not installed.", file=sys.stderr)
        print("Install with: pip install openai", file=sys.stderr)
        sys.exit(1)

    return OpenAI(base_url=base_url, api_key="not-needed")


# ============================================================================
# REFERENCE FILE LOADING
# ============================================================================

def get_skill_dir() -> Path:
    """Get the skill directory (parent of scripts/)."""
    return Path(__file__).parent.parent


def load_reference_file(tumor_type: str) -> str:
    """Load the appropriate reference file for the tumor type."""
    skill_dir = get_skill_dir()
    
    reference_map = {
        "pancreas": "diagnosis/exocrine_pancreas.md",
        "breast": "diagnosis/breast_invasive_carcinoma.md",
        "colorectal": "diagnosis/colorectal_resection.md",
        "gastric": "diagnosis/gastric_carcinoma.md",
        "stomach": "diagnosis/gastric_carcinoma.md",
    }
    
    filename = reference_map.get(tumor_type)
    if not filename:
        return ""
    
    ref_path = skill_dir / "references" / filename
    
    if ref_path.exists():
        return ref_path.read_text(encoding="utf-8")
    return ""


def load_staging_reference() -> str:
    """Load the TNM stage calculator reference file."""
    skill_dir = get_skill_dir()
    staging_path = skill_dir / "references" / "staging" / "tnm_stage_calculator.md"

    if staging_path.exists():
        return staging_path.read_text(encoding="utf-8")
    return ""


def load_skill_file() -> str:
    """Load the main SKILL.md file."""
    skill_dir = get_skill_dir()
    skill_path = skill_dir / "SKILL.md"
    
    if skill_path.exists():
        with open(skill_path, "r", encoding="utf-8") as f:
            return f.read()
    
    return ""


# ============================================================================
# LLM ANALYSIS
# ============================================================================

ANALYSIS_PROMPT = """You are a pathology report compliance checker. Analyze the following surgical pathology cancer report against CAP (College of American Pathologists) and ICCR (International Collaboration on Cancer Reporting) guidelines.

<skill_instructions>
{skill_content}
</skill_instructions>

<reference_guidelines>
{reference_content}
</reference_guidelines>

<staging_reference>
{staging_content}
</staging_reference>

<pathology_report>
{report_text}
</pathology_report>

Analyze this report and return a JSON object with the following structure:

```json
{{
  "tumor_type": "pancreas|breast|colorectal|gastric",
  "compliance_score": 0-100,
  "status": "COMPLIANT|INCOMPLETE - MINOR|INCOMPLETE - MAJOR|INCOMPLETE - CRITICAL",
  "gaps": [
    {{
      "element": "element name",
      "severity": "critical|major|minor",
      "status": "missing|empty",
      "guideline": "CAP + ICCR core",
      "required_for": "staging, prognosis, etc."
    }}
  ],
  "present_elements": [
    {{
      "element": "element name",
      "value": "extracted value"
    }}
  ],
  "staging": {{
    "pT": "extracted pT category or null",
    "pN": "extracted pN category or null",
    "pM": "extracted pM/M category or null",
    "reported_stage": "stage from report or null",
    "calculated_stage": "calculated stage based on pT/pN/pM or null",
    "stage_match": true|false|null,
    "stage_notes": "any notes about staging (discrepancy explanation, ypTNM, etc.)"
  }},
  "validation_issues": [
    {{
      "rule_name": "pT vs Size",
      "description": "description of the inconsistency",
      "found_value": "what was found in report",
      "expected_value": "what should be there",
      "severity": "error|warning",
      "note": "AJCC/guideline reference"
    }}
  ],
  "quality_metrics": {{
    "completeness_score": 0-100,
    "clarity_score": 0-100,
    "consistency_score": 0-100,
    "overall_quality": 0-100,
    "completeness_details": "brief explanation",
    "clarity_details": "brief explanation",
    "consistency_details": "brief explanation"
  }},
  "summary": "Brief narrative summary of findings"
}}
```

Important:
1. Detect tumor type from the report if not specified
2. Check ALL required elements from the reference guidelines
3. Distinguish between MISSING (field not present) and EMPTY (field label exists but no value)
4. Perform cross-validation checks (pT vs size, pN vs node count, margin vs R classification, node adequacy)
5. Calculate quality metrics based on completeness, clarity, and consistency
6. Use AJCC 8th edition staging criteria for validation
7. VERIFY TNM STAGE: Extract pT, pN, pM from report, calculate expected stage using staging_reference, compare with reported stage
8. If stage discrepancy found, add to validation_issues with severity "error"

Return ONLY the JSON object, no other text."""


def analyze_report_with_llm(
    client,
    report_text: str,
    tumor_type: Optional[str],
    skill_content: str,
    model: str,
    provider: str = "anthropic"
) -> Dict[str, Any]:
    """Analyze a single report using the specified LLM provider."""

    # Load appropriate reference file
    if tumor_type:
        reference_content = load_reference_file(tumor_type)
    else:
        # Load all references for auto-detection
        reference_content = "\n\n---\n\n".join([
            load_reference_file("pancreas"),
            load_reference_file("breast"),
            load_reference_file("colorectal"),
            load_reference_file("gastric"),
        ])

    # Load staging reference
    staging_content = load_staging_reference()

    prompt = ANALYSIS_PROMPT.format(
        skill_content=skill_content,
        reference_content=reference_content,
        staging_content=staging_content,
        report_text=report_text
    )

    try:
        # Call appropriate API
        if provider == "anthropic":
            response = client.messages.create(
                model=model,
                max_tokens=4096,
                messages=[{"role": "user", "content": prompt}]
            )
            response_text = response.content[0].text
        else:
            # OpenAI-compatible API (Ollama, LM Studio)
            response = client.chat.completions.create(
                model=model,
                max_tokens=4096,
                messages=[{"role": "user", "content": prompt}]
            )
            response_text = response.choices[0].message.content

        # Parse JSON from response
        # Handle potential markdown code blocks
        json_match = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', response_text)
        if json_match:
            json_str = json_match.group(1)
        else:
            json_str = response_text

        result = json.loads(json_str)
        result["_raw_response"] = response_text
        return result

    except json.JSONDecodeError as e:
        print(f"  Warning: Failed to parse JSON response: {e}")
        return {
            "error": f"JSON parse error: {e}",
            "tumor_type": tumor_type or "unknown",
            "compliance_score": 0,
            "status": "ERROR",
            "gaps": [],
            "present_elements": [],
            "validation_issues": [],
            "quality_metrics": None,
            "_raw_response": response_text if 'response_text' in locals() else ""
        }
    except Exception as e:
        print(f"  Warning: API error: {e}")
        return {
            "error": f"API error: {e}",
            "tumor_type": tumor_type or "unknown",
            "compliance_score": 0,
            "status": "ERROR",
            "gaps": [],
            "present_elements": [],
            "validation_issues": [],
            "quality_metrics": None,
            "_raw_response": ""
        }


def parse_llm_result(llm_result: Dict[str, Any], filename: str) -> ReportResult:
    """Convert LLM result dict to ReportResult dataclass."""
    
    # Parse gaps
    gaps = []
    for g in llm_result.get("gaps", []):
        gaps.append(GapInfo(
            element=g.get("element", ""),
            severity=g.get("severity", "minor"),
            status=g.get("status", "missing"),
            guideline=g.get("guideline", ""),
            required_for=g.get("required_for", "")
        ))
    
    # Parse validation issues
    validation_issues = []
    for v in llm_result.get("validation_issues", []):
        validation_issues.append(ValidationIssue(
            rule_name=v.get("rule_name", ""),
            description=v.get("description", ""),
            found_value=v.get("found_value", ""),
            expected_value=v.get("expected_value", ""),
            severity=v.get("severity", "warning"),
            note=v.get("note", "")
        ))
    
    # Parse quality metrics
    qm_data = llm_result.get("quality_metrics")
    quality_metrics = None
    if qm_data:
        quality_metrics = QualityMetrics(
            completeness_score=qm_data.get("completeness_score", 0),
            clarity_score=qm_data.get("clarity_score", 0),
            consistency_score=qm_data.get("consistency_score", 0),
            overall_quality=qm_data.get("overall_quality", 0),
            completeness_details=qm_data.get("completeness_details", ""),
            clarity_details=qm_data.get("clarity_details", ""),
            consistency_details=qm_data.get("consistency_details", "")
        )
    
    # Parse staging information
    staging_data = llm_result.get("staging")
    staging = None
    if staging_data:
        staging = StagingInfo(
            pT=staging_data.get("pT"),
            pN=staging_data.get("pN"),
            pM=staging_data.get("pM"),
            reported_stage=staging_data.get("reported_stage"),
            calculated_stage=staging_data.get("calculated_stage"),
            stage_match=staging_data.get("stage_match"),
            stage_notes=staging_data.get("stage_notes", "")
        )
    
    # Parse present elements
    present_elements = []
    for p in llm_result.get("present_elements", []):
        if isinstance(p, dict):
            element = p.get('element', '') or ''
            value = p.get('value', '') or ''
            present_elements.append(f"{element}: {str(value)[:50]}...")
        else:
            present_elements.append(str(p))
    
    # Count by severity
    critical_count = sum(1 for g in gaps if g.severity == "critical")
    major_count = sum(1 for g in gaps if g.severity == "major")
    minor_count = sum(1 for g in gaps if g.severity == "minor")
    missing_count = sum(1 for g in gaps if g.status == "missing")
    empty_count = sum(1 for g in gaps if g.status == "empty")

    # Handle tumor_type being a list or string
    tumor_type_raw = llm_result.get("tumor_type", "unknown")
    if isinstance(tumor_type_raw, list):
        tumor_type = tumor_type_raw[0] if tumor_type_raw else "unknown"
    else:
        tumor_type = str(tumor_type_raw) if tumor_type_raw else "unknown"

    return ReportResult(
        filename=filename,
        tumor_type=tumor_type,
        score=llm_result.get("compliance_score", 0),
        status=llm_result.get("status", "UNKNOWN"),
        critical_count=critical_count,
        major_count=major_count,
        minor_count=minor_count,
        missing_count=missing_count,
        empty_count=empty_count,
        present_count=len(present_elements),
        gaps=gaps,
        present_elements=present_elements,
        validation_issues=validation_issues,
        quality_metrics=quality_metrics,
        staging=staging,
        raw_analysis=llm_result.get("summary", "")
    )


# ============================================================================
# BATCH PROCESSING
# ============================================================================

def process_directory(
    input_dir: str,
    output_dir: str,
    tumor_type: Optional[str] = None,
    model: Optional[str] = None,
    provider: str = "anthropic",
    ollama_url: str = DEFAULT_OLLAMA_URL,
    lmstudio_url: str = DEFAULT_LMSTUDIO_URL,
    rate_limit_delay: float = 1.0
) -> List[ReportResult]:
    """Process all report files in a directory using the specified LLM provider."""

    # Create appropriate client and determine model
    if provider == "anthropic":
        client = create_anthropic_client()
        model = model or DEFAULT_ANTHROPIC_MODEL

    elif provider == "ollama":
        client = create_openai_compatible_client(ollama_url)
        if not model:
            model = auto_select_model(ollama_url, provider) or DEFAULT_OLLAMA_MODEL

    elif provider == "lmstudio":
        client = create_openai_compatible_client(lmstudio_url)
        if not model:
            model = auto_select_model(lmstudio_url, provider)
            if not model:
                print("Error: No models loaded in LM Studio.")
                print("Open LM Studio → Local Server tab → Load a model → Start Server")
                sys.exit(1)

    else:
        print(f"Error: Unknown provider '{provider}'")
        sys.exit(1)

    input_path = Path(input_dir)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Create subdirectory for individual reports
    individual_dir = output_path / "individual_reports"
    individual_dir.mkdir(exist_ok=True)

    # Load skill content once
    skill_content = load_skill_file()

    results = []

    # Find all text files
    report_files = list(input_path.glob("*.txt")) + list(input_path.glob("*.TXT"))

    print(f"Found {len(report_files)} report files to process...")
    print(f"Using provider: {provider}")
    print(f"Using model: {model}")
    print()

    for i, report_file in enumerate(report_files, 1):
        print(f"Processing [{i}/{len(report_files)}]: {report_file.name}")

        try:
            with open(report_file, "r", encoding="utf-8") as f:
                report_text = f.read()
        except UnicodeDecodeError:
            with open(report_file, "r", encoding="latin-1") as f:
                report_text = f.read()

        # Analyze with LLM
        llm_result = analyze_report_with_llm(
            client=client,
            report_text=report_text,
            tumor_type=tumor_type,
            skill_content=skill_content,
            model=model,
            provider=provider
        )

        # Parse result
        result = parse_llm_result(llm_result, report_file.name)
        results.append(result)

        # Save individual report
        individual_report = generate_individual_report(result)
        with open(individual_dir / f"{report_file.stem}_qa.txt", "w", encoding="utf-8") as f:
            f.write(individual_report)

        # Rate limiting (only for cloud providers)
        if provider == "anthropic" and i < len(report_files):
            time.sleep(rate_limit_delay)

    return results


# ============================================================================
# REPORT GENERATION
# ============================================================================

def generate_individual_report(result: ReportResult) -> str:
    """Generate a formatted QA report for a single result."""
    
    lines = [
        "═" * 65,
        "PATHOLOGY REPORT COMPLIANCE CHECK ",
        "═" * 65,
        "",
        f"FILENAME: {result.filename}",
        f"TUMOR TYPE: {result.tumor_type.upper()}",
        f"DATE CHECKED: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "─" * 65,
        "COMPLIANCE SCORE",
        "─" * 65,
        f"SCORE: {result.score}/100",
        f"STATUS: {result.status}",
        "",
        "Gap Summary:",
        f"  🔴 Critical: {result.critical_count} elements",
        f"  🟠 Major: {result.major_count} elements",
        f"  🟡 Minor: {result.minor_count} elements",
        "",
        "Field Status:",
        f"  ⬜ Missing fields: {result.missing_count}",
        f"  ⬚ Empty fields: {result.empty_count}",
        "",
    ]
    
    # Group gaps by severity
    critical_gaps = [g for g in result.gaps if g.severity == "critical"]
    major_gaps = [g for g in result.gaps if g.severity == "major"]
    minor_gaps = [g for g in result.gaps if g.severity == "minor"]
    
    if critical_gaps:
        lines.extend([
            "─" * 65,
            "CRITICAL GAPS (Severity: 🔴)",
            "─" * 65,
        ])
        for i, gap in enumerate(critical_gaps, 1):
            lines.extend([
                f"{i}. {gap.element.upper()}",
                f"   - Status: {gap.status.upper()}",
                f"   - Guideline: {gap.guideline}",
                f"   - Required for: {gap.required_for}",
                ""
            ])
    
    if major_gaps:
        lines.extend([
            "─" * 65,
            "MAJOR GAPS (Severity: 🟠)",
            "─" * 65,
        ])
        for i, gap in enumerate(major_gaps, 1):
            lines.extend([
                f"{i}. {gap.element.upper()}",
                f"   - Status: {gap.status.upper()}",
                f"   - Guideline: {gap.guideline}",
                ""
            ])
    
    if minor_gaps:
        lines.extend([
            "─" * 65,
            "MINOR GAPS (Severity: 🟡)",
            "─" * 65,
        ])
        for i, gap in enumerate(minor_gaps, 1):
            lines.extend([
                f"{i}. {gap.element.upper()}",
                f"   - Status: {gap.status.upper()}",
                ""
            ])
    
    lines.extend([
        "─" * 65,
        f"PRESENT ELEMENTS ({result.present_count} verified)",
        "─" * 65,
    ])
    for elem in result.present_elements[:10]:
        lines.append(f"✓ {elem}")
    if len(result.present_elements) > 10:
        lines.append(f"  ... and {len(result.present_elements) - 10} more")
    
    # Validation issues
    if result.validation_issues:
        lines.extend([
            "",
            "─" * 65,
            "CROSS-VALIDATION ISSUES",
            "─" * 65,
        ])
        
        errors = [v for v in result.validation_issues if v.severity == "error"]
        warnings = [v for v in result.validation_issues if v.severity == "warning"]
        
        if errors:
            lines.append("❌ ERRORS:")
            for issue in errors:
                lines.extend([
                    f"   {issue.rule_name}",
                    f"   - Found: {issue.found_value}",
                    f"   - Expected: {issue.expected_value}",
                    f"   - Note: {issue.note}",
                    ""
                ])
        
        if warnings:
            lines.append("⚠️ WARNINGS:")
            for issue in warnings:
                lines.extend([
                    f"   {issue.rule_name}",
                    f"   - Found: {issue.found_value}",
                    f"   - Expected: {issue.expected_value}",
                    f"   - Note: {issue.note}",
                    ""
                ])
    
    # Quality metrics
    if result.quality_metrics:
        qm = result.quality_metrics
        lines.extend([
            "",
            "─" * 65,
            "QUALITY METRICS",
            "─" * 65,
            f"Overall Quality Score: {qm.overall_quality}/100",
            "",
            f"  📊 Completeness: {qm.completeness_score}/100",
            f"     {qm.completeness_details}",
            "",
            f"  📝 Clarity: {qm.clarity_score}/100",
            f"     {qm.clarity_details}",
            "",
            f"  🔗 Consistency: {qm.consistency_score}/100",
            f"     {qm.consistency_details}",
        ])
    
    # TNM Staging verification
    if result.staging:
        stg = result.staging
        lines.extend([
            "",
            "─" * 65,
            "TNM STAGE VERIFICATION",
            "─" * 65,
            "Extracted from report:",
            f"  pT: {stg.pT or 'Not specified'}",
            f"  pN: {stg.pN or 'Not specified'}",
            f"  pM: {stg.pM or 'Not specified'}",
            "",
            f"Reported Stage: {stg.reported_stage or 'Not stated'}",
            f"Calculated Stage: {stg.calculated_stage or 'Cannot calculate'}",
            "",
        ])
        if stg.stage_match is True:
            lines.append("Status: ✓ STAGE MATCHES")
        elif stg.stage_match is False:
            lines.append("Status: ❌ STAGE DISCREPANCY")
            lines.append(f"  Expected: {stg.calculated_stage}")
            lines.append(f"  Reported: {stg.reported_stage}")
        else:
            lines.append("Status: ⚠️ Could not verify (missing TNM data)")
        
        if stg.stage_notes:
            lines.append(f"Notes: {stg.stage_notes}")
    
    # Summary
    if result.raw_analysis:
        lines.extend([
            "",
            "─" * 65,
            "ANALYSIS SUMMARY",
            "─" * 65,
            result.raw_analysis
        ])
    
    return "\n".join(lines)


def generate_summary_report(results: List[ReportResult]) -> str:
    """Generate overall summary statistics."""
    
    total = len(results)
    if total == 0:
        return "No reports processed."
    
    # Filter out errors
    valid_results = [r for r in results if r.status != "ERROR"]
    valid_count = len(valid_results)
    
    if valid_count == 0:
        return "All reports failed to process."
    
    compliant = sum(1 for r in valid_results if r.status == "COMPLIANT")
    minor = sum(1 for r in valid_results if r.status == "INCOMPLETE - MINOR")
    major = sum(1 for r in valid_results if r.status == "INCOMPLETE - MAJOR")
    critical = sum(1 for r in valid_results if r.status == "INCOMPLETE - CRITICAL")
    
    avg_score = sum(r.score or 0 for r in valid_results) / valid_count
    
    # Quality metrics averages
    qm_results = [r for r in valid_results if r.quality_metrics]
    if qm_results:
        avg_completeness = sum(r.quality_metrics.completeness_score for r in qm_results) / len(qm_results)
        avg_clarity = sum(r.quality_metrics.clarity_score for r in qm_results) / len(qm_results)
        avg_consistency = sum(r.quality_metrics.consistency_score for r in qm_results) / len(qm_results)
        avg_quality = sum(r.quality_metrics.overall_quality for r in qm_results) / len(qm_results)
    else:
        avg_completeness = avg_clarity = avg_consistency = avg_quality = 0
    
    # Validation issue counts
    total_errors = sum(len([v for v in r.validation_issues if v.severity == "error"]) for r in valid_results)
    total_warnings = sum(len([v for v in r.validation_issues if v.severity == "warning"]) for r in valid_results)
    reports_with_issues = sum(1 for r in valid_results if r.validation_issues)
    
    # Most common gaps
    gap_counts = {}
    for r in valid_results:
        for gap in r.gaps:
            gap_counts[gap.element] = gap_counts.get(gap.element, 0) + 1
    sorted_gaps = sorted(gap_counts.items(), key=lambda x: x[1], reverse=True)
    
    # Most common validation issues
    issue_counts = {}
    for r in valid_results:
        for issue in r.validation_issues:
            issue_counts[issue.rule_name] = issue_counts.get(issue.rule_name, 0) + 1
    sorted_issues = sorted(issue_counts.items(), key=lambda x: x[1], reverse=True)
    
    lines = [
        "═" * 65,
        "BATCH COMPLIANCE SUMMARY REPORT ",
        "═" * 65,
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "─" * 65,
        "OVERVIEW",
        "─" * 65,
        f"Total reports processed: {total}",
        f"Successfully analyzed: {valid_count}",
        f"Average compliance score: {avg_score:.1f}/100",
        "",
        "Compliance Distribution:",
        f"  ✅ Compliant (90-100):        {compliant:3d} ({100*compliant/valid_count:.1f}%)",
        f"  🟡 Incomplete - Minor (70-89): {minor:3d} ({100*minor/valid_count:.1f}%)",
        f"  🟠 Incomplete - Major (50-69): {major:3d} ({100*major/valid_count:.1f}%)",
        f"  🔴 Incomplete - Critical (<50): {critical:3d} ({100*critical/valid_count:.1f}%)",
        "",
        "─" * 65,
        "QUALITY METRICS (Averages)",
        "─" * 65,
        f"  📊 Overall Quality:  {avg_quality:.1f}/100",
        f"  📋 Completeness:     {avg_completeness:.1f}/100",
        f"  📝 Clarity:          {avg_clarity:.1f}/100",
        f"  🔗 Consistency:      {avg_consistency:.1f}/100",
        "",
        "─" * 65,
        "CROSS-VALIDATION SUMMARY",
        "─" * 65,
        f"  Reports with issues: {reports_with_issues}/{valid_count} ({100*reports_with_issues/valid_count:.1f}%)",
        f"  Total errors:        {total_errors}",
        f"  Total warnings:      {total_warnings}",
    ]
    
    if sorted_issues:
        lines.append("")
        lines.append("  Most common issues:")
        for issue_name, count in sorted_issues[:5]:
            lines.append(f"    - {issue_name}: {count} reports")
    
    lines.extend([
        "",
        "─" * 65,
        "MOST COMMON GAPS (Top 10)",
        "─" * 65,
    ])
    
    for element, count in sorted_gaps[:10]:
        pct = 100 * count / valid_count
        lines.append(f"  {element:30s}: {count:3d} reports ({pct:.1f}%)")
    
    lines.extend([
        "",
        "─" * 65,
        "SCORE DISTRIBUTION",
        "─" * 65,
    ])
    
    brackets = [(90, 100), (70, 89), (50, 69), (0, 49)]
    for low, high in brackets:
        count = sum(1 for r in valid_results if low <= r.score <= high)
        bar = "█" * (count * 40 // valid_count) if valid_count > 0 else ""
        lines.append(f"  {low:2d}-{high:3d}: {bar} {count}")
    
    return "\n".join(lines)


# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def export_to_csv(results: List[ReportResult], output_path: Path):
    """Export results to CSV."""
    
    csv_file = output_path / "compliance_data.csv"
    
    with open(csv_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        
        writer.writerow([
            "Filename", "Tumor Type", "Compliance Score", "Status",
            "Critical Gaps", "Major Gaps", "Minor Gaps",
            "Missing Fields", "Empty Fields", "Present Elements",
            "Overall Quality", "Completeness", "Clarity", "Consistency",
            "pT", "pN", "pM", "Reported Stage", "Calculated Stage", "Stage Match",
            "Validation Errors", "Validation Warnings",
            "Gap List", "Validation Issues"
        ])
        
        for r in results:
            gap_list = "; ".join([g.element for g in r.gaps])
            
            qm = r.quality_metrics
            quality = qm.overall_quality if qm else ""
            completeness = qm.completeness_score if qm else ""
            clarity = qm.clarity_score if qm else ""
            consistency = qm.consistency_score if qm else ""
            
            # Staging info
            stg = r.staging
            pT = stg.pT if stg else ""
            pN = stg.pN if stg else ""
            pM = stg.pM if stg else ""
            reported_stage = stg.reported_stage if stg else ""
            calculated_stage = stg.calculated_stage if stg else ""
            stage_match = "Yes" if stg and stg.stage_match is True else ("No" if stg and stg.stage_match is False else "")
            
            errors = len([v for v in r.validation_issues if v.severity == "error"])
            warnings = len([v for v in r.validation_issues if v.severity == "warning"])
            issue_list = "; ".join([f"{v.rule_name}: {v.found_value}" for v in r.validation_issues])
            
            writer.writerow([
                r.filename, r.tumor_type, r.score, r.status,
                r.critical_count, r.major_count, r.minor_count,
                r.missing_count, r.empty_count, r.present_count,
                quality, completeness, clarity, consistency,
                pT, pN, pM, reported_stage, calculated_stage, stage_match,
                errors, warnings,
                gap_list, issue_list
            ])
    
    print(f"CSV exported: {csv_file}")


def export_to_excel(results: List[ReportResult], output_path: Path):
    """Export results to Excel with multiple sheets."""
    
    if not EXCEL_AVAILABLE:
        print("Skipping Excel export (openpyxl not installed)")
        return
    
    xlsx_file = output_path / "compliance_data.xlsx"
    wb = Workbook()
    
    valid_results = [r for r in results if r.status != "ERROR"]
    total = len(valid_results)
    
    if total == 0:
        print("No valid results to export")
        return
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    avg_score = sum(r.score or 0 for r in valid_results) / total
    
    qm_results = [r for r in valid_results if r.quality_metrics]
    if qm_results:
        avg_quality = sum(r.quality_metrics.overall_quality for r in qm_results) / len(qm_results)
        avg_completeness = sum(r.quality_metrics.completeness_score for r in qm_results) / len(qm_results)
        avg_clarity = sum(r.quality_metrics.clarity_score for r in qm_results) / len(qm_results)
        avg_consistency = sum(r.quality_metrics.consistency_score for r in qm_results) / len(qm_results)
    else:
        avg_quality = avg_completeness = avg_clarity = avg_consistency = 0
    
    summary_data = [
        ["Batch Compliance Summary ", ""],
        ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M')],
        ["", ""],
        ["Total Reports", total],
        ["", ""],
        ["COMPLIANCE SCORES", ""],
        ["Average Compliance Score", f"{avg_score:.1f}"],
        ["", ""],
        ["Status", "Count", "Percentage"],
        ["Compliant", sum(1 for r in valid_results if r.status == "COMPLIANT"), ""],
        ["Incomplete - Minor", sum(1 for r in valid_results if r.status == "INCOMPLETE - MINOR"), ""],
        ["Incomplete - Major", sum(1 for r in valid_results if r.status == "INCOMPLETE - MAJOR"), ""],
        ["Incomplete - Critical", sum(1 for r in valid_results if r.status == "INCOMPLETE - CRITICAL"), ""],
        ["", ""],
        ["QUALITY METRICS", ""],
        ["Average Overall Quality", f"{avg_quality:.1f}"],
        ["Average Completeness", f"{avg_completeness:.1f}"],
        ["Average Clarity", f"{avg_clarity:.1f}"],
        ["Average Consistency", f"{avg_consistency:.1f}"],
    ]
    
    for row in summary_data:
        ws_summary.append(row)
    
    ws_summary["A1"].font = Font(bold=True, size=14)
    
    # Sheet 2: All Results
    ws_results = wb.create_sheet("All Results")
    
    headers = ["Filename", "Tumor Type", "Compliance", "Status", 
               "Critical", "Major", "Minor", "Missing", "Empty", "Present",
               "Quality", "Completeness", "Clarity", "Consistency",
               "Val Errors", "Val Warnings"]
    ws_results.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_results.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for r in valid_results:
        qm = r.quality_metrics
        errors = len([v for v in r.validation_issues if v.severity == "error"])
        warnings = len([v for v in r.validation_issues if v.severity == "warning"])
        
        ws_results.append([
            r.filename, r.tumor_type, r.score, r.status,
            r.critical_count, r.major_count, r.minor_count,
            r.missing_count, r.empty_count, r.present_count,
            qm.overall_quality if qm else "", 
            qm.completeness_score if qm else "",
            qm.clarity_score if qm else "",
            qm.consistency_score if qm else "",
            errors, warnings
        ])
    
    # Sheet 3: Gap Analysis
    ws_gaps = wb.create_sheet("Gap Analysis")
    
    gap_headers = ["Element", "Count", "Percentage", "Severity"]
    ws_gaps.append(gap_headers)
    
    for col, header in enumerate(gap_headers, 1):
        cell = ws_gaps.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    gap_counts = {}
    gap_severity = {}
    for r in valid_results:
        for gap in r.gaps:
            gap_counts[gap.element] = gap_counts.get(gap.element, 0) + 1
            gap_severity[gap.element] = gap.severity
    
    for element, count in sorted(gap_counts.items(), key=lambda x: x[1], reverse=True):
        pct = 100 * count / total
        ws_gaps.append([element, count, f"{pct:.1f}%", gap_severity.get(element, "")])
    
    # Sheet 4: Validation Issues
    ws_validation = wb.create_sheet("Validation Issues")
    
    val_headers = ["Filename", "Rule", "Severity", "Found Value", "Expected Value", "Note"]
    ws_validation.append(val_headers)
    
    for col, header in enumerate(val_headers, 1):
        cell = ws_validation.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for r in valid_results:
        for issue in r.validation_issues:
            ws_validation.append([
                r.filename, issue.rule_name, issue.severity,
                issue.found_value, issue.expected_value, issue.note
            ])
    
    # Sheet 5: Quality Details
    ws_quality = wb.create_sheet("Quality Details")
    
    qual_headers = ["Filename", "Overall", "Completeness", "Clarity", "Consistency",
                    "Completeness Details", "Clarity Details", "Consistency Details"]
    ws_quality.append(qual_headers)
    
    for col, header in enumerate(qual_headers, 1):
        cell = ws_quality.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for r in valid_results:
        qm = r.quality_metrics
        if qm:
            ws_quality.append([
                r.filename, qm.overall_quality, qm.completeness_score, 
                qm.clarity_score, qm.consistency_score,
                qm.completeness_details, qm.clarity_details, qm.consistency_details
            ])
    
    wb.save(xlsx_file)
    print(f"Excel exported: {xlsx_file}")


def save_trend_data(results: List[ReportResult], output_path: Path):
    """Save trend data for historical tracking."""
    
    valid_results = [r for r in results if r.status != "ERROR"]
    if not valid_results:
        return
    
    total = len(valid_results)
    
    # Calculate averages
    avg_compliance = sum(r.score for r in valid_results) / total
    
    qm_results = [r for r in valid_results if r.quality_metrics]
    if qm_results:
        avg_completeness = sum(r.quality_metrics.completeness_score for r in qm_results) / len(qm_results)
        avg_clarity = sum(r.quality_metrics.clarity_score for r in qm_results) / len(qm_results)
        avg_consistency = sum(r.quality_metrics.consistency_score for r in qm_results) / len(qm_results)
        avg_quality = sum(r.quality_metrics.overall_quality for r in qm_results) / len(qm_results)
    else:
        avg_completeness = avg_clarity = avg_consistency = avg_quality = 0
    
    # Common gaps
    gap_counts = {}
    for r in valid_results:
        for gap in r.gaps:
            gap_counts[gap.element] = gap_counts.get(gap.element, 0) + 1
    common_gaps = sorted(gap_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Common issues
    issue_counts = {}
    for r in valid_results:
        for issue in r.validation_issues:
            issue_counts[issue.rule_name] = issue_counts.get(issue.rule_name, 0) + 1
    common_issues = sorted(issue_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    
    trend = {
        "date": datetime.now().strftime('%Y-%m-%d'),
        "report_count": total,
        "avg_compliance": round(avg_compliance, 1),
        "avg_completeness": round(avg_completeness, 1),
        "avg_clarity": round(avg_clarity, 1),
        "avg_consistency": round(avg_consistency, 1),
        "avg_quality": round(avg_quality, 1),
        "common_gaps": common_gaps,
        "common_issues": common_issues
    }
    
    # Load existing history
    trend_file = output_path / "trend_history.json"
    history = []
    
    if trend_file.exists():
        try:
            with open(trend_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        except (json.JSONDecodeError, IOError):
            history = []
    
    history.append(trend)
    
    with open(trend_file, 'w', encoding='utf-8') as f:
        json.dump(history, f, indent=2, ensure_ascii=False)
    
    print(f"Trend data saved: {trend_file}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Batch process pathology reports for CAP/ICCR compliance",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process with Claude (default)
  python batch_checker.py /path/to/reports /path/to/output

  # Process with local Ollama
  python batch_checker.py /reports /output --provider ollama -m qwen2.5:3b

  # Process with LM Studio
  python batch_checker.py /reports /output --provider lmstudio

  # List available models
  python batch_checker.py --list-models --provider ollama

  # Specify tumor type
  python batch_checker.py /reports /output --tumor-type breast
"""
    )

    parser.add_argument(
        "input_dir",
        nargs="?",
        help="Directory containing report text files (.txt)"
    )

    parser.add_argument(
        "output_dir",
        nargs="?",
        help="Directory to save results"
    )

    parser.add_argument(
        "--provider", "-p",
        type=str,
        default=os.environ.get("LLM_PROVIDER", "anthropic"),
        choices=["anthropic", "ollama", "lmstudio"],
        help="LLM provider (default: anthropic)"
    )

    parser.add_argument(
        "--model", "-m",
        type=str,
        default=os.environ.get("OLLAMA_MODEL") or os.environ.get("LMSTUDIO_MODEL"),
        help="Model name (auto-detected for local providers if not specified)"
    )

    parser.add_argument(
        "--list-models",
        action="store_true",
        help="List available models for the selected provider and exit"
    )

    parser.add_argument(
        "--ollama-url",
        type=str,
        default=os.environ.get("OLLAMA_URL", DEFAULT_OLLAMA_URL),
        help=f"Ollama API base URL (default: {DEFAULT_OLLAMA_URL})"
    )

    parser.add_argument(
        "--lmstudio-url",
        type=str,
        default=os.environ.get("LMSTUDIO_URL", DEFAULT_LMSTUDIO_URL),
        help=f"LM Studio API base URL (default: {DEFAULT_LMSTUDIO_URL})"
    )

    parser.add_argument(
        "--tumor-type", "-t",
        type=str,
        choices=["breast", "colorectal", "pancreas", "gastric"],
        help="Tumor type hint (auto-detected if not specified)"
    )

    args = parser.parse_args()

    # Handle --list-models
    if args.list_models:
        if args.provider == "anthropic":
            print("Model listing not supported for Anthropic. Available models:")
            print(f"  {DEFAULT_ANTHROPIC_MODEL}")
            sys.exit(0)

        base_url = args.lmstudio_url if args.provider == "lmstudio" else args.ollama_url
        models = list_available_models(base_url, args.provider)

        if not models:
            print(f"No models found. Is {args.provider} running?", file=sys.stderr)
            if args.provider == "ollama":
                print("Start with: ollama serve", file=sys.stderr)
            else:
                print("Open LM Studio → Local Server → Load a model → Start Server", file=sys.stderr)
            sys.exit(1)

        print(f"Available models ({args.provider}):")
        for model_id in models:
            print(f"  {model_id}")
        sys.exit(0)

    # Validate required arguments
    if not args.input_dir or not args.output_dir:
        parser.print_help()
        sys.exit(1)

    print(f"Input directory: {args.input_dir}")
    print(f"Output directory: {args.output_dir}")
    print(f"Provider: {args.provider}")
    print(f"Tumor type: {args.tumor_type or 'auto-detect'}")
    print()

    # Process reports
    results = process_directory(
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        tumor_type=args.tumor_type,
        model=args.model,
        provider=args.provider,
        ollama_url=args.ollama_url,
        lmstudio_url=args.lmstudio_url
    )

    if not results:
        print("No reports found to process.")
        sys.exit(1)

    output_path = Path(args.output_dir)

    # Generate summary report
    summary = generate_summary_report(results)
    summary_file = output_path / "summary_report.txt"
    with open(summary_file, "w", encoding="utf-8") as f:
        f.write(summary)
    print(f"\nSummary report: {summary_file}")

    # Print summary
    print("\n" + summary)

    # Export
    export_to_csv(results, output_path)
    export_to_excel(results, output_path)
    save_trend_data(results, output_path)

    print(f"\nProcessing complete. Results saved to: {args.output_dir}")


if __name__ == "__main__":
    main()
