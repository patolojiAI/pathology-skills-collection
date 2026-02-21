"""
Excel output generation for batch pathology report processing.

Creates multi-sheet workbooks with summary, detailed results, gap analysis, etc.
"""

from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


def export_results_to_excel(results: List[Dict[str, Any]], output_path: Path, skill_name: str = ""):
    """
    Generate multi-sheet Excel workbook with comprehensive results.

    Args:
        results: List of analysis results (dicts with keys: filename, compliance_score, gaps, etc.)
        output_path: Where to save the Excel file
        skill_name: Name of the skill used (for metadata)

    Sheets created:
        1. Summary - Overall statistics
        2. All Results - Per-report compliance data
        3. Gap Analysis - Most common missing elements
        4. Details - Full analysis for each report
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create sheets
    create_summary_sheet(wb, results, skill_name)
    create_results_sheet(wb, results)
    create_gap_analysis_sheet(wb, results)
    create_details_sheet(wb, results)

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    logger.info(f"Excel report saved to {output_path}")


def create_summary_sheet(wb, results: List[Dict], skill_name: str):
    """Create summary statistics sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet("Summary", 0)

    # Title
    ws['A1'] = f"Pathology Report Analysis Summary - {skill_name}"
    ws['A1'].font = Font(size=14, bold=True)

    # Metadata
    ws['A3'] = "Generated:"
    ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws['A4'] = "Total Reports:"
    ws['B4'] = len(results)

    # Calculate statistics
    if results:
        scores = [r.get('compliance_score', 0) for r in results if 'compliance_score' in r]
        avg_score = sum(scores) / len(scores) if scores else 0

        compliant = sum(1 for s in scores if s >= 90)
        minor_incomplete = sum(1 for s in scores if 70 <= s < 90)
        major_incomplete = sum(1 for s in scores if 50 <= s < 70)
        critical_incomplete = sum(1 for s in scores if s < 50)

        ws['A5'] = "Average Compliance Score:"
        ws['B5'] = f"{avg_score:.1f}"

        # Compliance distribution
        ws['A7'] = "Compliance Distribution"
        ws['A7'].font = Font(bold=True)

        ws['A8'] = "✅ Compliant (90-100):"
        ws['B8'] = compliant
        ws['C8'] = f"{compliant/len(scores)*100:.1f}%" if scores else "0%"

        ws['A9'] = "🟡 Minor Issues (70-89):"
        ws['B9'] = minor_incomplete
        ws['C9'] = f"{minor_incomplete/len(scores)*100:.1f}%" if scores else "0%"

        ws['A10'] = "🟠 Major Issues (50-69):"
        ws['B10'] = major_incomplete
        ws['C10'] = f"{major_incomplete/len(scores)*100:.1f}%" if scores else "0%"

        ws['A11'] = "🔴 Critical Issues (<50):"
        ws['B11'] = critical_incomplete
        ws['C11'] = f"{critical_incomplete/len(scores)*100:.1f}%" if scores else "0%"

        # Most common gaps
        all_gaps = []
        for r in results:
            if 'gaps' in r and r['gaps']:
                if isinstance(r['gaps'], list):
                    all_gaps.extend(r['gaps'])
                elif isinstance(r['gaps'], str):
                    all_gaps.extend([g.strip() for g in r['gaps'].split(',') if g.strip()])

        if all_gaps:
            from collections import Counter
            gap_counts = Counter(all_gaps)
            top_gaps = gap_counts.most_common(5)

            ws['A13'] = "Top 5 Most Common Gaps"
            ws['A13'].font = Font(bold=True)

            for idx, (gap, count) in enumerate(top_gaps, start=14):
                ws[f'A{idx}'] = gap
                ws[f'B{idx}'] = count
                ws[f'C{idx}'] = f"{count/len(results)*100:.1f}%"

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def create_results_sheet(wb, results: List[Dict]):
    """Create detailed results sheet with all reports."""
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("All Results")

    # Headers
    headers = ['Filename', 'Compliance Score', 'Status', 'Missing Elements', 'Gaps Count', 'Notes']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data rows
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result.get('filename', ''))

        score = result.get('compliance_score', 0)
        ws.cell(row=row_idx, column=2, value=score)

        # Status based on score
        if score >= 90:
            status = "✅ Compliant"
        elif score >= 70:
            status = "🟡 Minor Issues"
        elif score >= 50:
            status = "🟠 Major Issues"
        else:
            status = "🔴 Critical Issues"

        ws.cell(row=row_idx, column=3, value=status)

        # Gaps
        gaps = result.get('gaps', [])
        if isinstance(gaps, list):
            ws.cell(row=row_idx, column=4, value=', '.join(gaps[:3]))  # Show first 3
            ws.cell(row=row_idx, column=5, value=len(gaps))
        elif isinstance(gaps, str):
            ws.cell(row=row_idx, column=4, value=gaps[:100])
            ws.cell(row=row_idx, column=5, value=gaps.count(',') + 1 if gaps else 0)

        ws.cell(row=row_idx, column=6, value=result.get('notes', ''))

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 40


def create_gap_analysis_sheet(wb, results: List[Dict]):
    """Create gap analysis sheet showing most common missing elements."""
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("Gap Analysis")

    # Headers
    headers = ['Missing Element', 'Frequency', 'Percentage', 'Severity']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Collect all gaps
    all_gaps = []
    for r in results:
        if 'gaps' in r and r['gaps']:
            if isinstance(r['gaps'], list):
                all_gaps.extend(r['gaps'])
            elif isinstance(r['gaps'], str):
                all_gaps.extend([g.strip() for g in r['gaps'].split(',') if g.strip()])

    if all_gaps:
        from collections import Counter
        gap_counts = Counter(all_gaps)

        # Sort by frequency
        for row_idx, (gap, count) in enumerate(gap_counts.most_common(), start=2):
            ws.cell(row=row_idx, column=1, value=gap)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=f"{count/len(results)*100:.1f}%")

            # Infer severity from gap name (heuristic)
            gap_lower = gap.lower()
            if any(kw in gap_lower for kw in ['pt', 'pn', 'margin', 'grade', 'type']):
                severity = "🔴 CRITICAL"
            elif any(kw in gap_lower for kw in ['lvi', 'pni', 'size', 'node']):
                severity = "🟠 MAJOR"
            else:
                severity = "🟡 MINOR"

            ws.cell(row=row_idx, column=4, value=severity)

    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18


def create_details_sheet(wb, results: List[Dict]):
    """Create detailed sheet with full analysis text for each report."""
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet("Details")

    # Headers
    headers = ['Filename', 'Full Analysis', 'Recommendations']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data rows
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result.get('filename', ''))
        ws.cell(row=row_idx, column=2, value=result.get('analysis', 'N/A'))
        ws.cell(row=row_idx, column=3, value=result.get('recommendations', 'N/A'))

        # Wrap text for readability
        ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 50

    # Row heights
    for row in range(2, len(results) + 2):
        ws.row_dimensions[row].height = 100


def create_simple_excel(results: List[Dict], output_path: Path):
    """
    Create simple single-sheet Excel output (lightweight alternative).

    Useful for quick exports without detailed analysis.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl required. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Headers
    headers = ['Filename', 'Score', 'Status', 'Gaps']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result.get('filename', ''))
        ws.cell(row=row_idx, column=2, value=result.get('compliance_score', 0))

        score = result.get('compliance_score', 0)
        status = "✅" if score >= 90 else "🟡" if score >= 70 else "🟠" if score >= 50 else "🔴"
        ws.cell(row=row_idx, column=3, value=status)

        gaps = result.get('gaps', [])
        if isinstance(gaps, list):
            ws.cell(row=row_idx, column=4, value=', '.join(gaps))
        else:
            ws.cell(row=row_idx, column=4, value=str(gaps))

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    logger.info(f"Simple Excel saved to {output_path}")
